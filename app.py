from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from wtforms import Form, StringField, FloatField, IntegerField, TextAreaField
from wtforms.validators import DataRequired, NumberRange
from flask_mail import Mail, Message
from xhtml2pdf import pisa
from dotenv import load_dotenv
import json
import os
import io
import base64

app = Flask(__name__)
load_dotenv()

# Configuration
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///quotes.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Email Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_USERNAME')

mail = Mail(app)
db = SQLAlchemy(app)

# ========== DATABASE MODELS ==========

class Quote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quote_no = db.Column(db.String(50), unique=True, nullable=False)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    due_date = db.Column(db.DateTime, nullable=False)
    customer_name = db.Column(db.String(200), nullable=False)
    customer_address = db.Column(db.Text, nullable=False)
    header_text = db.Column(db.String(500), default="")
    total_amount = db.Column(db.Float, default=0.0)
    
    discount_type = db.Column(db.String(20), default="none")
    discount_value = db.Column(db.Float, default=0.0)
    discount_amount = db.Column(db.Float, default=0.0)
    apply_gst = db.Column(db.Boolean, default=True)
    
    cleaning_charges = db.Column(db.Boolean, default=False)
    cleaning_qty = db.Column(db.Float, default=0)
    cleaning_unit = db.Column(db.String(20), default="SQFT")
    cleaning_price = db.Column(db.Float, default=0)
    cleaning_amount = db.Column(db.Float, default=0)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship('QuoteItem', backref='quote', lazy=True, cascade='all, delete-orphan')

class QuoteItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quote_id = db.Column(db.Integer, db.ForeignKey('quote.id'), nullable=False)
    description = db.Column(db.String(500), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(20), default="SQFT")
    price = db.Column(db.Float, nullable=False)
    amount = db.Column(db.Float, nullable=False)

class ItemTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(500))
    default_price = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(20), default="SQFT")

class Contact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    address = db.Column(db.Text)
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ========== WORK ORDER MODEL ==========
class WorkOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    work_order_no = db.Column(db.String(50), unique=True, nullable=False)
    quote_id = db.Column(db.Integer, db.ForeignKey('quote.id'), nullable=False)
    quote = db.relationship('Quote', backref=db.backref('work_orders', cascade='all, delete-orphan'))
    
    date = db.Column(db.DateTime, default=datetime.utcnow)
    start_date = db.Column(db.DateTime, nullable=True)
    expected_end_date = db.Column(db.DateTime, nullable=True)
    
    customer_name = db.Column(db.String(200), nullable=False)
    customer_address = db.Column(db.Text, nullable=False)
    
    payment_term_1_label = db.Column(db.String(100), default="Advance Payment")
    payment_term_1_percentage = db.Column(db.Float, default=50.0)
    payment_term_2_label = db.Column(db.String(100), default="After Interior Painting")
    payment_term_2_percentage = db.Column(db.Float, default=30.0)
    payment_term_3_label = db.Column(db.String(100), default="After Exterior Painting")
    payment_term_3_percentage = db.Column(db.Float, default=15.0)
    payment_term_4_label = db.Column(db.String(100), default="After Final Completion")
    payment_term_4_percentage = db.Column(db.Float, default=5.0)
    
    terms_conditions = db.Column(db.Text, default="")
    status = db.Column(db.String(50), default="In Progress")
    notes = db.Column(db.Text, default="")
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def generate_work_order_no(self):
        year = datetime.utcnow().year
        last_wo = WorkOrder.query.filter(WorkOrder.work_order_no.like(f'WO-{year}-%')).order_by(WorkOrder.id.desc()).first()
        if last_wo:
            last_num = int(last_wo.work_order_no.split('-')[2])
            new_num = last_num + 1
        else:
            new_num = 1
        self.work_order_no = f"WO-{year}-{new_num:03d}"
    
    def get_payment_terms_list(self):
        painting_total = self.quote.total_amount or 0
        cleaning_total = self.quote.cleaning_amount or 0
        subtotal = painting_total + cleaning_total
        discount_amt = self.quote.discount_amount or 0
        after_discount = subtotal - discount_amt
        grand_total = after_discount * 1.18 if self.quote.apply_gst else after_discount
        
        terms = []
        if self.payment_term_1_percentage > 0:
            terms.append({
                'label': self.payment_term_1_label,
                'percentage': self.payment_term_1_percentage,
                'amount': (grand_total * self.payment_term_1_percentage) / 100
            })
        if self.payment_term_2_percentage > 0:
            terms.append({
                'label': self.payment_term_2_label,
                'percentage': self.payment_term_2_percentage,
                'amount': (grand_total * self.payment_term_2_percentage) / 100
            })
        if self.payment_term_3_percentage > 0:
            terms.append({
                'label': self.payment_term_3_label,
                'percentage': self.payment_term_3_percentage,
                'amount': (grand_total * self.payment_term_3_percentage) / 100
            })
        if self.payment_term_4_percentage > 0:
            terms.append({
                'label': self.payment_term_4_label,
                'percentage': self.payment_term_4_percentage,
                'amount': (grand_total * self.payment_term_4_percentage) / 100
            })
        return terms

# ========== FORMS ==========

class QuoteForm(Form):
    customer_name = StringField('Customer Name', validators=[DataRequired()])
    customer_address = TextAreaField('Address', validators=[DataRequired()])
    header_text = StringField('Header Text')
    due_date_offset = IntegerField('Due Date (Days)', default=90, validators=[NumberRange(min=1, max=365)])

# ========== HELPER FUNCTIONS ==========

def calculate_grand_total(quote):
    subtotal = quote.total_amount + (quote.cleaning_amount or 0)
    after_discount = subtotal - (quote.discount_amount or 0)
    if quote.apply_gst:
        return after_discount * 1.18
    return after_discount

app.jinja_env.globals.update(calculate_grand_total=calculate_grand_total)

# ========== ROUTES ==========

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    quotes = Quote.query.order_by(Quote.created_at.desc()).limit(10).all()
    total_quotes = Quote.query.count()
    total_amount = db.session.query(db.func.sum(Quote.total_amount)).scalar() or 0
    contacts_count = Contact.query.count()
    
    stats = {
        'total_quotes': total_quotes,
        'total_amount': total_amount,
        'recent_quotes': quotes
    }
    return render_template('dashboard.html', stats=stats, contacts_count=contacts_count)

@app.route('/quotes')
def quotes_list():
    quotes = Quote.query.order_by(Quote.created_at.desc()).all()
    return render_template('quotes_list.html', quotes=quotes)

# ========== DATA EXPORT ROUTES ==========

@app.route('/export-all-data')
def export_all_data():
    """Export all data - Quotes, Work Orders, Customers"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    wb = Workbook()
    
    # ===== Style Definitions =====
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== 1. QUOTES SHEET =====
    ws_quotes = wb.active
    ws_quotes.title = "Quotes"
    
    # Headers
    quote_headers = ['Quote No', 'Date', 'Due Date', 'Customer', 'Address', 
                     'Total Amount', 'Discount', 'GST', 'Grand Total']
    ws_quotes.append(quote_headers)
    
    # Style headers
    for col in range(1, len(quote_headers) + 1):
        cell = ws_quotes.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    quotes = Quote.query.all()
    for quote in quotes:
        subtotal = (quote.total_amount or 0) + (quote.cleaning_amount or 0)
        discount_amt = quote.discount_amount or 0
        after_discount = subtotal - discount_amt
        grand_total = after_discount * 1.18 if quote.apply_gst else after_discount
        
        ws_quotes.append([
            quote.quote_no,
            quote.date.strftime('%d/%m/%Y') if quote.date else 'N/A',
            quote.due_date.strftime('%d/%m/%Y') if quote.due_date else 'N/A',
            quote.customer_name,
            quote.customer_address[:100] if quote.customer_address else 'N/A',
            f"₹{subtotal:,.2f}",
            f"₹{discount_amt:,.2f}" if discount_amt > 0 else '₹0.00',
            f"₹{(after_discount * 0.18):,.2f}" if quote.apply_gst else '₹0.00',
            f"₹{grand_total:,.2f}"
        ])
    
    # Auto-adjust column widths
    for col in range(1, len(quote_headers) + 1):
        ws_quotes.column_dimensions[chr(64 + col)].width = 20
    
    # ===== 2. WORK ORDERS SHEET =====
    ws_wo = wb.create_sheet("Work Orders")
    
    wo_headers = ['Work Order No', 'Quote No', 'Customer', 'Date', 'Start Date', 
                  'End Date', 'Status', 'Total Amount', 'Payment Terms']
    ws_wo.append(wo_headers)
    
    # Style headers
    for col in range(1, len(wo_headers) + 1):
        cell = ws_wo.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    work_orders = WorkOrder.query.all()
    for wo in work_orders:
        subtotal = (wo.quote.total_amount or 0) + (wo.quote.cleaning_amount or 0)
        discount_amt = wo.quote.discount_amount or 0
        after_discount = subtotal - discount_amt
        grand_total = after_discount * 1.18 if wo.quote.apply_gst else after_discount
        
        ws_wo.append([
            wo.work_order_no,
            wo.quote.quote_no,
            wo.customer_name,
            wo.date.strftime('%d/%m/%Y') if wo.date else 'N/A',
            wo.start_date.strftime('%d/%m/%Y') if wo.start_date else 'N/A',
            wo.expected_end_date.strftime('%d/%m/%Y') if wo.expected_end_date else 'N/A',
            wo.status,
            f"₹{grand_total:,.2f}",
            f"{wo.payment_term_1_label} ({wo.payment_term_1_percentage}%), {wo.payment_term_2_label} ({wo.payment_term_2_percentage}%)"
        ])
    
    for col in range(1, len(wo_headers) + 1):
        ws_wo.column_dimensions[chr(64 + col)].width = 20
    
    # ===== 3. CUSTOMERS SHEET =====
    ws_customers = wb.create_sheet("Customers")
    
    customer_headers = ['Name', 'Address', 'Phone', 'Email', 'Created Date', 'Total Quotes', 'Total Amount']
    ws_customers.append(customer_headers)
    
    # Style headers
    for col in range(1, len(customer_headers) + 1):
        cell = ws_customers.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    customers = Contact.query.all()
    for customer in customers:
        # Get customer's quotes
        customer_quotes = Quote.query.filter_by(customer_name=customer.name).all()
        total_quotes = len(customer_quotes)
        total_amount = sum(q.total_amount for q in customer_quotes)
        
        ws_customers.append([
            customer.name,
            customer.address[:100] if customer.address else 'N/A',
            customer.phone if customer.phone else 'N/A',
            customer.email if customer.email else 'N/A',
            customer.created_at.strftime('%d/%m/%Y') if customer.created_at else 'N/A',
            total_quotes,
            f"₹{total_amount:,.2f}" if total_amount > 0 else '₹0.00'
        ])
    
    for col in range(1, len(customer_headers) + 1):
        ws_customers.column_dimensions[chr(64 + col)].width = 20
    
    # ===== 4. ITEMS SHEET =====
    ws_items = wb.create_sheet("Items")
    
    item_headers = ['Description', 'Quantity', 'Unit', 'Price', 'Amount', 'Quote No', 'Customer']
    ws_items.append(item_headers)
    
    # Style headers
    for col in range(1, len(item_headers) + 1):
        cell = ws_items.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    quote_items = QuoteItem.query.all()
    for item in quote_items:
        ws_items.append([
            item.description[:100] if item.description else 'N/A',
            item.quantity,
            item.unit if item.unit else 'SQFT',
            f"₹{item.price:,.2f}",
            f"₹{item.amount:,.2f}",
            item.quote.quote_no if item.quote else 'N/A',
            item.quote.customer_name if item.quote else 'N/A'
        ])
    
    for col in range(1, len(item_headers) + 1):
        ws_items.column_dimensions[chr(64 + col)].width = 20
    
    # ===== SAVE TO BYTES =====
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # ===== GET TODAY'S DATE =====
    today = datetime.now().strftime('%d-%m-%Y')
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=f'SHALOM_Data_Export_{today}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========== WORK ORDER ROUTES ==========

@app.route('/work-orders')
def work_orders_list():
    work_orders = WorkOrder.query.order_by(WorkOrder.created_at.desc()).all()
    return render_template('work_orders_list.html', work_orders=work_orders)

@app.route('/work-order/<int:work_order_id>')
def view_work_order(work_order_id):
    work_order = WorkOrder.query.get_or_404(work_order_id)
    return render_template('view_work_order.html', work_order=work_order)

# ========== CREATE WORK ORDER ROUTE ==========

@app.route('/quote/<int:quote_id>/create-work-order', methods=['GET', 'POST'])
def create_work_order(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    
    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('expected_end_date')
        
        start_date = None
        expected_end_date = None
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        if end_date_str:
            expected_end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        work_order = WorkOrder(
            quote_id=quote.id,
            customer_name=quote.customer_name,
            customer_address=quote.customer_address,
            start_date=start_date,
            expected_end_date=expected_end_date,
            status=request.form.get('status', 'In Progress'),
            notes=request.form.get('notes', ''),
            terms_conditions=request.form.get('terms_conditions', ''),
            payment_term_1_label=request.form.get('payment_term_1_label', 'Advance Payment'),
            payment_term_1_percentage=float(request.form.get('payment_term_1_percentage', 50)),
            payment_term_2_label=request.form.get('payment_term_2_label', 'After Interior Painting'),
            payment_term_2_percentage=float(request.form.get('payment_term_2_percentage', 30)),
            payment_term_3_label=request.form.get('payment_term_3_label', 'After Exterior Painting'),
            payment_term_3_percentage=float(request.form.get('payment_term_3_percentage', 15)),
            payment_term_4_label=request.form.get('payment_term_4_label', 'After Final Completion'),
            payment_term_4_percentage=float(request.form.get('payment_term_4_percentage', 5))
        )
        
        work_order.generate_work_order_no()
        
        db.session.add(work_order)
        db.session.commit()
        
        flash(f'Work Order {work_order.work_order_no} created successfully!', 'success')
        return redirect(url_for('view_work_order', work_order_id=work_order.id))
    
    return render_template('create_work_order.html', quote=quote)

# ========== WORK ORDER PDF EXPORT (SIMPLIFIED VERSION) ==========

@app.route('/work-order/<int:work_order_id>/export-pdf')
def export_work_order_pdf(work_order_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, Image, HRFlowable, PageBreak, KeepTogether)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    import io

    work_order = WorkOrder.query.get_or_404(work_order_id)
    quote = work_order.quote

    # ========== FONT SETUP (Rupee symbol; falls back to "Rs." if the
    # font file isn't deployed on the server — that's what happened in
    # your last render, so this is working as intended) ==========
    FONT_REGULAR = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'
    RUPEE = 'Rs. '

    font_candidates = [
        os.path.join(app.root_path, 'static', 'fonts', 'DejaVuSans.ttf'),
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    bold_candidates = [
        os.path.join(app.root_path, 'static', 'fonts', 'DejaVuSans-Bold.ttf'),
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    ]
    reg_path = next((p for p in font_candidates if os.path.exists(p)), None)
    bold_path = next((p for p in bold_candidates if os.path.exists(p)), None)

    if reg_path and bold_path:
        try:
            if 'DejaVuSans' not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont('DejaVuSans', reg_path))
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_path))
            FONT_REGULAR = 'DejaVuSans'
            FONT_BOLD = 'DejaVuSans-Bold'
            RUPEE = '\u20b9'
        except Exception:
            pass

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4,
                             topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    CONTENT_WIDTH = A4[0] - doc.leftMargin - doc.rightMargin  # ~17cm on A4
    half_width = CONTENT_WIDTH / 2

    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9, leading=13, fontName=FONT_REGULAR)
    bold_style = ParagraphStyle('Bold', parent=normal_style, fontName=FONT_BOLD)
    bold_right_style = ParagraphStyle('BoldRight', parent=bold_style, alignment=TA_RIGHT)
    center_style = ParagraphStyle('Center', parent=normal_style, alignment=TA_CENTER)
    right_style = ParagraphStyle('Right', parent=normal_style, alignment=TA_RIGHT)
    table_header_style = ParagraphStyle('TH', parent=normal_style, fontName=FONT_BOLD, fontSize=9, textColor=colors.white, alignment=TA_CENTER)

    # ========== HEADER — logo only, no company name text ==========
    logo_path = os.path.join(app.root_path, 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=8 * cm, height=2 * cm)
            logo_img.hAlign = 'CENTER'
            story.append(logo_img)
        except Exception:
            pass
    story.append(Spacer(1, 0.3 * cm))

    # ========== WORK ORDER HEADER ==========
    wo_title_and_info = []
    wo_title_and_info.append(Paragraph("<b>WORK ORDER</b>", ParagraphStyle('WOTitle', parent=styles['Normal'], fontSize=16, alignment=TA_CENTER, textColor=colors.HexColor('#1e3c72'), spaceAfter=10, fontName=FONT_BOLD)))

    wo_info_data = [
        [Paragraph(f"<b>Work Order No:</b> {work_order.work_order_no}", normal_style),
         Paragraph(f"<b>Date:</b> {work_order.date.strftime('%d/%m/%Y')}", normal_style)],
        [Paragraph(f"<b>Quote No:</b> {quote.quote_no}", normal_style),
         Paragraph(f"<b>Status:</b> {work_order.status}", bold_style)],
    ]
    wo_info_table = Table(wo_info_data, colWidths=[half_width, half_width])
    wo_info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
    ]))
    wo_title_and_info.append(wo_info_table)
    wo_title_and_info.append(Spacer(1, 0.3 * cm))

    dates_data = [
        [Paragraph(f"<b>Start Date:</b> {work_order.start_date.strftime('%d/%m/%Y') if work_order.start_date else 'Not specified'}", normal_style),
         Paragraph(f"<b>Expected End Date:</b> {work_order.expected_end_date.strftime('%d/%m/%Y') if work_order.expected_end_date else 'Not specified'}", normal_style)]
    ]
    dates_table = Table(dates_data, colWidths=[half_width, half_width])
    dates_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
    ]))
    wo_title_and_info.append(dates_table)
    wo_title_and_info.append(Spacer(1, 0.3 * cm))

    # Keep the WO title + both info strips together so they never split
    # across a page boundary awkwardly.
    story.append(KeepTogether(wo_title_and_info))

    # ========== CUSTOMER DETAILS ==========
    customer_block = [
        Paragraph("<b>Customer Details</b>", bold_style),
    ]
    customer_address = work_order.customer_address.replace('\n', '<br/>') if work_order.customer_address else 'Address not provided'
    customer_block.append(Paragraph(f"{work_order.customer_name}<br/>{customer_address}", normal_style))
    customer_block.append(Spacer(1, 0.5 * cm))
    story.append(KeepTogether(customer_block))

    # ========== SCOPE OF WORK ==========
    scope_block = [
        Paragraph("<b>Scope of Work</b>", bold_style),
        Spacer(1, 0.2 * cm),
    ]

    items_data = [
        [Paragraph("Description", table_header_style),
         Paragraph("Qty", table_header_style),
         Paragraph("Unit", table_header_style),
         Paragraph(f"Amount ({RUPEE})", table_header_style)]
    ]

    for item in quote.items:
        items_data.append([
            Paragraph(item.description, normal_style),
            Paragraph(str(int(item.quantity)), center_style),
            Paragraph(item.unit or 'SQFT', center_style),
            Paragraph(f"{item.amount:,.2f}", right_style)
        ])

    if quote.cleaning_charges and quote.cleaning_amount > 0:
        items_data.append([
            Paragraph("<b>Cleaning and Handling Charges</b><br/><font size=8>Post-construction cleaning</font>", normal_style),
            Paragraph(str(int(quote.cleaning_qty or 0)), center_style),
            Paragraph(quote.cleaning_unit or 'SQFT', center_style),
            Paragraph(f"<b>{quote.cleaning_amount:,.2f}</b>", right_style)
        ])

    col_desc = CONTENT_WIDTH * 0.59
    col_qty = CONTENT_WIDTH * 0.11
    col_unit = CONTENT_WIDTH * 0.11
    col_amount = CONTENT_WIDTH * 0.19

    items_table = Table(items_data, colWidths=[col_desc, col_qty, col_unit, col_amount])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3c72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ALIGN', (1, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1e3c72')),
        ('LINEBELOW', (0, 1), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
    ]))
    scope_block.append(items_table)
    story.append(KeepTogether(scope_block))
    story.append(Spacer(1, 0.3 * cm))

    # ========== TOTALS — rebuilt as its own aligned table ==========
    # Previously the sign (+/-) sat directly in front of "Rs." with no
    # fixed column, so each row's digits started at a different
    # horizontal position — that's the "zig-zag" you saw. Now sign and
    # amount each get their own column, both right-aligned, so every
    # number lines up vertically underneath the one above it.
    painting_total = quote.total_amount or 0
    cleaning_total = quote.cleaning_amount or 0
    subtotal = painting_total + cleaning_total
    discount_amt = quote.discount_amount or 0
    after_discount = subtotal - discount_amt
    gst_amt = (after_discount * 0.18) if quote.apply_gst else 0
    grand_total = after_discount + gst_amt

    label_style = ParagraphStyle('TotalLabel', parent=bold_style, alignment=TA_RIGHT)
    sign_style = ParagraphStyle('TotalSign', parent=bold_style, alignment=TA_RIGHT)
    value_style = ParagraphStyle('TotalValue', parent=bold_style, alignment=TA_RIGHT)
    red_sign_style = ParagraphStyle('RedSign', parent=sign_style, textColor=colors.red)
    red_value_style = ParagraphStyle('RedValue', parent=value_style, textColor=colors.red)
    grand_style = ParagraphStyle('GrandTotal', parent=bold_style, fontSize=11, alignment=TA_RIGHT, textColor=colors.HexColor('#1e3c72'))

    total_rows = []
    total_rows.append([Paragraph('Subtotal', label_style), '', Paragraph(f'{RUPEE}{subtotal:,.2f}', value_style)])

    if discount_amt > 0:
        total_rows.append([Paragraph(f'Discount ({quote.discount_value:.2f}%)', label_style),
                            Paragraph('-', red_sign_style),
                            Paragraph(f'{RUPEE}{discount_amt:,.2f}', red_value_style)])
        total_rows.append([Paragraph('After Discount', label_style), '',
                            Paragraph(f'{RUPEE}{after_discount:,.2f}', value_style)])

    if quote.apply_gst:
        total_rows.append([Paragraph('GST (18%)', label_style),
                            Paragraph('+', sign_style),
                            Paragraph(f'{RUPEE}{gst_amt:,.2f}', value_style)])

    total_rows.append([Paragraph('GRAND TOTAL', grand_style), '',
                        Paragraph(f'{RUPEE}{grand_total:,.2f}', grand_style)])

    # Reuses the exact same column widths as the items table
    # (desc+qty | unit | amount) so the totals block's right edge and
    # amount column line up perfectly under the items table above it.
    total_table = Table(
        total_rows,
        colWidths=[col_desc + col_qty, col_unit, col_amount]
    )
    last_row = len(total_rows) - 1
    total_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -2), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -2), 4),
        ('LINEABOVE', (0, last_row), (-1, last_row), 1, colors.HexColor('#1e3c72')),
        ('TOPPADDING', (0, last_row), (-1, last_row), 8),
        ('BOTTOMPADDING', (0, last_row), (-1, last_row), 8),
        ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#e8f0fe')),
    ]))
    story.append(total_table)
    story.append(Spacer(1, 0.5 * cm))

    # ========== PAYMENT SCHEDULE ==========
    payment_block = [
        HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3c72'), spaceBefore=5, spaceAfter=5),
        Paragraph("<b>Payment Schedule</b>", bold_style),
        Spacer(1, 0.2 * cm),
    ]

    payment_terms = work_order.get_payment_terms_list()

    if payment_terms:
        payment_data = [
            [Paragraph("<b>#</b>", table_header_style),
             Paragraph("<b>Milestone</b>", table_header_style),
             Paragraph("<b>Percentage</b>", table_header_style),
             Paragraph(f"<b>Amount ({RUPEE})</b>", table_header_style)]
        ]

        for idx, term in enumerate(payment_terms, 1):
            payment_data.append([
                Paragraph(str(idx), center_style),
                Paragraph(term['label'], normal_style),
                Paragraph(f"{term['percentage']:.1f}%", center_style),
                Paragraph(f"{RUPEE}{term['amount']:,.2f}", right_style)
            ])

        payment_data.append([
            "",
    Paragraph("<b>Total</b>", bold_style),
            "",
    Paragraph(f"<b>{RUPEE}{grand_total:,.2f}</b>", bold_right_style)
])

        col_num = CONTENT_WIDTH * 0.115
        col_milestone = CONTENT_WIDTH * 0.385
        col_pct = CONTENT_WIDTH * 0.23
        col_amt = CONTENT_WIDTH * 0.27

        payment_table = Table(payment_data, colWidths=[col_num, col_milestone, col_pct, col_amt])
        payment_table.setStyle(TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3c72')),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
    ('FONTSIZE', (0, 0), (-1, 0), 9),
    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ('TOPPADDING', (0, 0), (-1, 0), 8),
    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
    ('ALIGN', (2, 1), (2, -1), 'CENTER'),
    ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
    ('FONTSIZE', (0, 1), (-1, -1), 9),
    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
    ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1e3c72')),
    ('LINEBELOW', (0, 1), (-1, -2), 0.5, colors.HexColor('#e0e0e0')),
    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f0fe')),
    ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#1e3c72')),
]))
        payment_block.append(payment_table)

    story.append(KeepTogether(payment_block))
    story.append(Spacer(1, 0.3 * cm))

    # ========== PAGE BREAK BEFORE THE SIGNING PAGE ==========
    # Terms, notes, and the signature block belong together as the part
    # the customer actually signs — force them onto a fresh page instead
    # of letting them start wherever the Payment Schedule happens to end.
    story.append(PageBreak())

    signing_block = []

    signing_block.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3c72'), spaceBefore=5, spaceAfter=5))
    signing_block.append(Paragraph("<b>Terms and Conditions</b>", bold_style))

    terms_lines = work_order.terms_conditions.split('\n') if work_order.terms_conditions else []
    if terms_lines:
        for line in terms_lines:
            if line.strip():
                signing_block.append(Paragraph(f"• {line.strip()}", normal_style))

    signing_block.append(Spacer(1, 0.5 * cm))

    if work_order.notes:
        signing_block.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3c72'), spaceBefore=5, spaceAfter=5))
        signing_block.append(Paragraph("<b>Notes</b>", bold_style))
        signing_block.append(Paragraph(work_order.notes, normal_style))
        signing_block.append(Spacer(1, 0.5 * cm))

    signing_block.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3c72'), spaceBefore=10, spaceAfter=5))

    sig_data = [
        [Paragraph("<b>Client Signature</b>", center_style),
         Paragraph("<b>Authorized Signatory</b>", center_style)],
        [Paragraph("<br/><br/><br/>", normal_style),
         Paragraph("<br/><br/><br/>", normal_style)],
        ["_________________________", "_________________________"]
    ]
    sig_table = Table(sig_data, colWidths=[half_width, half_width])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
    ]))
    signing_block.append(sig_table)
    signing_block.append(Spacer(1, 0.5 * cm))

    story.append(KeepTogether(signing_block))

    # ========== FOOTER ==========
    story.append(Spacer(1, 0.8 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3c72'), spaceBefore=5, spaceAfter=5))

    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], alignment=TA_CENTER, fontSize=8, textColor=colors.HexColor('#666666'), fontName=FONT_REGULAR)
    story.append(Paragraph("SHALOM ENTERPRISES | Paint by Shalom | Professional Painting Services", footer_style))

    doc.build(story)
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f'WorkOrder_{work_order.work_order_no}_{work_order.customer_name}.pdf',
        mimetype='application/pdf'
    )
# ========== OTHER ROUTES ==========

@app.route('/create-quote', methods=['GET', 'POST'])
def create_quote():
    form = QuoteForm(request.form)
    
    if request.method == 'POST' and form.validate():
        due_date = datetime.utcnow() + timedelta(days=form.due_date_offset.data)
        
        year = datetime.utcnow().year
        last_quote = Quote.query.filter(Quote.quote_no.like(f'{year}-%')).order_by(Quote.id.desc()).first()
        if last_quote:
            last_num = int(last_quote.quote_no.split('-')[1])
            new_num = last_num + 1
        else:
            new_num = 1
        quote_no = f"{year}-{new_num:03d}"
        
        items_data = json.loads(request.form.get('items_data', '[]'))
        subtotal = sum(float(item['amount']) for item in items_data)
        
        discount_type = request.form.get('discount_type', 'none')
        discount_value = float(request.form.get('discount_value', 0))
        apply_gst = request.form.get('apply_gst') == 'true'
        
        discount_amount = 0
        if discount_type == 'percentage':
            discount_amount = (subtotal * discount_value) / 100
        elif discount_type == 'fixed':
            discount_amount = discount_value
        
        cleaning_data = json.loads(request.form.get('cleaning_data', '{"enabled":false}'))
        cleaning_enabled = cleaning_data.get('enabled', False)
        cleaning_qty = cleaning_data.get('qty', 0)
        cleaning_price = cleaning_data.get('price', 0)
        cleaning_amount = cleaning_data.get('amount', 0) if cleaning_enabled else 0
        
        save_as_contact = request.form.get('save_as_contact') == 'on'
        contact_id = request.form.get('contact_id')
        
        if save_as_contact and not contact_id:
            existing_contact = Contact.query.filter_by(name=form.customer_name.data).first()
            if not existing_contact:
                new_contact = Contact(
                    name=form.customer_name.data,
                    address=form.customer_address.data,
                    phone=request.form.get('customer_phone', ''),
                    email=request.form.get('customer_email', '')
                )
                db.session.add(new_contact)
                db.session.commit()
        
        quote = Quote(
            quote_no=quote_no,
            date=datetime.utcnow(),
            due_date=due_date,
            customer_name=form.customer_name.data,
            customer_address=form.customer_address.data,
            header_text=form.header_text.data,
            total_amount=subtotal,
            discount_type=discount_type,
            discount_value=discount_value,
            discount_amount=discount_amount,
            apply_gst=apply_gst,
            cleaning_charges=cleaning_enabled,
            cleaning_qty=cleaning_qty,
            cleaning_price=cleaning_price,
            cleaning_amount=cleaning_amount
        )
        
        db.session.add(quote)
        db.session.commit()
        
        for item_data in items_data:
            item = QuoteItem(
                quote_id=quote.id,
                description=item_data['description'],
                quantity=float(item_data['quantity']),
                unit=item_data['unit'],
                price=float(item_data['price']),
                amount=float(item_data['amount'])
            )
            db.session.add(item)
        
        db.session.commit()
        
        flash(f'Quote {quote_no} created successfully!', 'success')
        return redirect(url_for('view_quote', quote_id=quote.id))
    
    return render_template('create_quote.html', form=form)

@app.route('/quote/<int:quote_id>')
def view_quote(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    return render_template('view_quote.html', quote=quote)

@app.route('/quote/<int:quote_id>/edit', methods=['GET', 'POST'])
def edit_quote(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    
    if request.method == 'POST':
        quote.customer_name = request.form.get('customer_name')
        quote.customer_address = request.form.get('customer_address')
        quote.header_text = request.form.get('header_text')
        quote.discount_type = request.form.get('discount_type', 'none')
        quote.discount_value = float(request.form.get('discount_value', 0))
        quote.apply_gst = request.form.get('apply_gst') == 'true'
        
        items_data = json.loads(request.form.get('items_data', '[]'))
        subtotal = sum(float(item['amount']) for item in items_data)
        
        if quote.discount_type == 'percentage':
            quote.discount_amount = (subtotal * quote.discount_value) / 100
        elif quote.discount_type == 'fixed':
            quote.discount_amount = quote.discount_value
        else:
            quote.discount_amount = 0
        
        quote.total_amount = subtotal
        
        for item in quote.items:
            db.session.delete(item)
        
        for item_data in items_data:
            item = QuoteItem(
                quote_id=quote.id,
                description=item_data['description'],
                quantity=float(item_data['quantity']),
                unit=item_data['unit'],
                price=float(item_data['price']),
                amount=float(item_data['amount'])
            )
            db.session.add(item)
        
        db.session.commit()
        flash('Quote updated successfully!', 'success')
        return redirect(url_for('view_quote', quote_id=quote.id))
    
    return render_template('edit_quote.html', quote=quote)

@app.route('/quote/<int:quote_id>/delete')
def delete_quote(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    try:
        db.session.delete(quote)
        db.session.commit()
        flash('Quote deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Could not delete quote: {str(e)}', 'danger')
    return redirect(url_for('quotes_list'))

@app.route('/quote/<int:quote_id>/export-pdf')
def export_quote_pdf(quote_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    import io

    quote = Quote.query.get_or_404(quote_id)

    # ========== FONT SETUP (fixes the "black square" issue) ==========
    # Helvetica has no glyph for the Rupee sign, so ReportLab draws a solid
    # box instead. DejaVu Sans does include it. If the font file isn't
    # found (e.g. not deployed to the server), we fall back to plain
    # Helvetica and print "Rs." instead of the symbol so nothing breaks.
    FONT_REGULAR = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'
    RUPEE = 'Rs. '  # safe fallback

    font_candidates = [
        os.path.join(app.root_path, 'static', 'fonts', 'DejaVuSans.ttf'),
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    bold_candidates = [
        os.path.join(app.root_path, 'static', 'fonts', 'DejaVuSans-Bold.ttf'),
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    ]
    reg_path = next((p for p in font_candidates if os.path.exists(p)), None)
    bold_path = next((p for p in bold_candidates if os.path.exists(p)), None)

    if reg_path and bold_path:
        try:
            pdfmetrics.registerFont(TTFont('DejaVuSans', reg_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_path))
            FONT_REGULAR = 'DejaVuSans'
            FONT_BOLD = 'DejaVuSans-Bold'
            RUPEE = '\u20b9'  # real ₹ glyph, safe now that the font supports it
        except Exception:
            pass  # keep the Helvetica / "Rs." fallback

    # Create PDF in memory
    pdf_buffer = io.BytesIO()

    # ========== PAGE / WIDTH SETUP ==========
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4,
                             topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    # Usable width = page width minus left/right margins. Every table below
    # is sized (or padded) to exactly this width so their edges line up.
    CONTENT_WIDTH = A4[0] - doc.leftMargin - doc.rightMargin  # ~17cm on A4

    story = []
    styles = getSampleStyleSheet()

    # ========== CUSTOM STYLES ==========
    title_style = ParagraphStyle(
        'CompanyTitle', parent=styles['Heading1'], fontSize=20,
        textColor=colors.HexColor('#1e3c72'), spaceAfter=2, fontName=FONT_BOLD
    )
    subtitle_style = ParagraphStyle(
        'CompanySubtitle', parent=styles['Normal'], fontSize=10,
        textColor=colors.HexColor('#2a5298'), spaceAfter=15, fontName=FONT_REGULAR
    )
    normal_style = ParagraphStyle(
        'Normal', parent=styles['Normal'], fontSize=9, leading=13, fontName=FONT_REGULAR
    )
    bold_style = ParagraphStyle('Bold', parent=normal_style, fontName=FONT_BOLD)
    right_style = ParagraphStyle('Right', parent=normal_style, alignment=TA_RIGHT)
    center_style = ParagraphStyle('Center', parent=normal_style, alignment=TA_CENTER)
    table_header_style = ParagraphStyle(
        'TableHeader', parent=normal_style, fontName=FONT_BOLD, fontSize=9,
        textColor=colors.white, alignment=TA_CENTER
    )

    # ========== HEADER ==========
    logo_path = os.path.join(app.root_path, 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=13 * cm, height=4 * cm)
            img.hAlign = 'CENTER'
            story.append(img)
        except Exception:
            pass

    # Quote Info — widths now sum to CONTENT_WIDTH so this lines up with
    # every other block on the page.
    quote_info_data = [
        [Paragraph(f"<b>Quote No:</b> {quote.quote_no}", normal_style),
         Paragraph(f"<b>Date:</b> {quote.date.strftime('%d/%m/%Y') if quote.date else 'N/A'}", right_style)],
        ["", Paragraph(f"<b>Due Date:</b> {quote.due_date.strftime('%d/%m/%Y') if quote.due_date else 'N/A'}", right_style)]
    ]
    info_table = Table(quote_info_data, colWidths=[CONTENT_WIDTH * 0.5, CONTENT_WIDTH * 0.5])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#1e3c72')),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5 * cm))

    # ========== CUSTOMER INFO ==========
    customer_name = quote.customer_name or 'Customer'
    customer_address = quote.customer_address.replace('\n', '<br/>') if quote.customer_address else 'Address not provided'

    story.append(Paragraph("<b>Bill To:</b>", bold_style))
    story.append(Paragraph(f"{customer_name}<br/>{customer_address}", normal_style))
    story.append(Spacer(1, 0.5 * cm))

    # ========== ITEMS TABLE ==========
    # Column widths now sum to CONTENT_WIDTH (~17cm on A4) instead of 18cm,
    # so the table no longer gets auto-squeezed by ReportLab.
    col_desc = CONTENT_WIDTH * 0.50
    col_qty = CONTENT_WIDTH * 0.08
    col_unit = CONTENT_WIDTH * 0.08
    col_price = CONTENT_WIDTH * 0.17
    col_amount = CONTENT_WIDTH * 0.17

    table_data = [
        [Paragraph("Description", table_header_style),
         Paragraph("Qty", table_header_style),
         Paragraph("Unit", table_header_style),
         Paragraph(f"Price ({RUPEE})", table_header_style),
         Paragraph(f"Amount ({RUPEE})", table_header_style)]
    ]

    for item in quote.items:
        table_data.append([
            Paragraph(item.description, normal_style),
            Paragraph(str(int(item.quantity)), center_style),
            Paragraph(item.unit or 'SQFT', center_style),
            Paragraph(f"{item.price:,.2f}", right_style),
            Paragraph(f"{item.amount:,.2f}", right_style)
        ])

    if quote.cleaning_charges and quote.cleaning_amount > 0:
        table_data.append([
            Paragraph("<b>Cleaning and Handling Charges</b><br/><font size=8>Post-construction cleaning</font>", normal_style),
            Paragraph(str(int(quote.cleaning_qty or 0)), center_style),
            Paragraph(quote.cleaning_unit or 'SQFT', center_style),
            Paragraph(f"{quote.cleaning_price:,.2f}", right_style),
            Paragraph(f"<b>{quote.cleaning_amount:,.2f}</b>", right_style)
        ])

    table = Table(table_data, colWidths=[col_desc, col_qty, col_unit, col_price, col_amount])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3c72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ALIGN', (1, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1e3c72')),
        ('LINEBELOW', (0, 1), (-1, -2), 0.5, colors.HexColor('#e0e0e0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.5 * cm))

    # ========== CALCULATIONS ==========
    painting_total = quote.total_amount or 0
    cleaning_total = quote.cleaning_amount or 0
    subtotal = painting_total + cleaning_total
    discount_amt = quote.discount_amount or 0
    after_discount = subtotal - discount_amt
    gst_amt = (after_discount * 0.18) if quote.apply_gst else 0
    grand_total = after_discount + gst_amt

    # ========== TOTALS TABLE ==========
    # Same trick as the items table: this is built as a 3-column table
    # (spacer | label | value) whose widths sum to CONTENT_WIDTH, so its
    # right edge lands exactly under the items table's right edge instead
    # of floating as its own narrow, disconnected block.
    label_style = ParagraphStyle('TotalLabel', parent=normal_style, alignment=TA_RIGHT)
    value_style = ParagraphStyle('TotalValue', parent=normal_style, alignment=TA_RIGHT)
    grand_label_style = ParagraphStyle(
        'GrandLabel', parent=normal_style, alignment=TA_RIGHT,
        fontName=FONT_BOLD, fontSize=12, textColor=colors.HexColor('#1e3c72')
    )
    grand_value_style = ParagraphStyle(
        'GrandValue', parent=normal_style, alignment=TA_RIGHT,
        fontName=FONT_BOLD, fontSize=12, textColor=colors.HexColor('#1e3c72')
    )

    total_rows = []
    total_rows.append(['', Paragraph('Subtotal:', label_style), Paragraph(f'{RUPEE} {subtotal:,.2f}', value_style)])
    if discount_amt > 0:
        total_rows.append(['', Paragraph(f'Discount ({quote.discount_value:.2f}%):', label_style),
                            Paragraph(f'- {RUPEE} {discount_amt:,.2f}', value_style)])
        total_rows.append(['', Paragraph('After Discount:', label_style),
                            Paragraph(f'{RUPEE} {after_discount:,.2f}', value_style)])
    if quote.apply_gst:
        total_rows.append(['', Paragraph('GST (18%):', label_style), Paragraph(f'+ {RUPEE} {gst_amt:,.2f}', value_style)])
    total_rows.append(['', Paragraph('GRAND TOTAL:', grand_label_style),
                        Paragraph(f'{RUPEE} {grand_total:,.2f}', grand_value_style)])

    spacer_col = CONTENT_WIDTH - 7 * cm  # leaves 4cm label + 3cm value
    total_table = Table(total_rows, colWidths=[spacer_col, 4 * cm, 3 * cm])
    last_row = len(total_rows) - 1
    total_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ('TOPPADDING', (1, 0), (2, -2), 4),
        ('BOTTOMPADDING', (1, 0), (2, -2), 4),
        ('LINEABOVE', (1, last_row), (2, last_row), 1.2, colors.HexColor('#1e3c72')),
        ('TOPPADDING', (1, last_row), (2, last_row), 10),
        ('BOTTOMPADDING', (1, last_row), (2, last_row), 10),
        ('BACKGROUND', (1, last_row), (2, last_row), colors.HexColor('#eef2fa')),
    ]))
    story.append(total_table)
    story.append(Spacer(1, 0.8 * cm))

    # ========== MESSAGE ==========
    message_style = ParagraphStyle(
        'Message', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10,
        textColor=colors.grey, fontName=FONT_REGULAR, spaceAfter=10
    )
    story.append(Paragraph("I hope that this offer is what you are looking for. I am available to provide any additional information you may require.", message_style))
    story.append(Spacer(1, 0.2 * cm))

    # ========== SIGNATURE ==========
    story.append(Paragraph("<b>Best regards,</b>", ParagraphStyle('Regards', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=11, spaceAfter=5, fontName=FONT_REGULAR)))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("<b>Mathan Kumar</b>", ParagraphStyle('Signature', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=10, spaceAfter=2, fontName=FONT_REGULAR)))
    story.append(Paragraph("<b>For Shalom Enterprises</b>", ParagraphStyle('SigTitle', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=9, textColor=colors.grey, spaceAfter=10, fontName=FONT_REGULAR)))
    story.append(Spacer(1, 0.5 * cm))

    # ========== APPROVAL ==========
   # story.append(Paragraph("<b>Read and Approved</b>", ParagraphStyle('ApprovalTitle', parent=styles['Normal'], fontSize=11, spaceAfter=5, fontName=FONT_REGULAR)))
   # story.append(Paragraph(quote.customer_name or 'Customer', normal_style))
   # story.append(Spacer(1, 0.5 * cm))

   # approval_col = CONTENT_WIDTH / 3
   # approval_data = [
   #     [Paragraph("<b>Name</b>", normal_style),
   #      Paragraph("<b>Location and Date</b>", normal_style),
   #      Paragraph("<b>Signature</b>", normal_style)],
   ##     ['', '', '']
  #  ]
  #  approval_table = Table(approval_data, colWidths=[approval_col, approval_col, approval_col])
  #  approval_table.setStyle(TableStyle([
  #      ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
   #     ('FONTSIZE', (0, 0), (-1, 0), 9),
    #    ('TOPPADDING', (0, 0), (-1, 0), 8),
     ##  ('LINEBELOW', (0, 1), (-1, 1), 1, colors.black),
       # ('TOPPADDING', (0, 1), (-1, 1), 15),
        #('BOTTOMPADDING', (0, 1), (-1, 1), 5),
        #('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    #]))
    #story.append(approval_table)

    # ========== FOOTER ==========
    story.append(Spacer(1, 0.8 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3c72'), spaceBefore=5, spaceAfter=5))
    footer_style = ParagraphStyle(
        'Footer', parent=styles['Normal'], alignment=TA_CENTER, fontSize=8,
        textColor=colors.HexColor('#999999'), fontName=FONT_REGULAR
    )
    story.append(Paragraph("SHALOM ENTERPRISES | Paint by Shalom | Professional Painting Services", footer_style))

    # ========== BUILD PDF ==========
    doc.build(story)
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f'Estimate_{quote.quote_no}_{quote.customer_name}.pdf',
        mimetype='application/pdf'
    )

@app.route('/contacts')
def contacts():
    contacts = Contact.query.all()
    return render_template('contacts.html', contacts=contacts)

@app.route('/add-contact', methods=['POST'])
def add_contact():
    name = request.form.get('name')
    existing_contact = Contact.query.filter_by(name=name).first()
    if existing_contact:
        flash('Contact with this name already exists!', 'warning')
        return redirect(url_for('contacts'))
    
    contact = Contact(
        name=name,
        address=request.form.get('address'),
        phone=request.form.get('phone'),
        email=request.form.get('email')
    )
    db.session.add(contact)
    db.session.commit()
    flash('Contact added successfully!', 'success')
    return redirect(url_for('contacts'))

@app.route('/delete-contact/<int:contact_id>')
def delete_contact(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    db.session.delete(contact)
    db.session.commit()
    flash('Contact deleted successfully!', 'success')
    return redirect(url_for('contacts'))

@app.route('/items')
def items():
    items = ItemTemplate.query.all()
    return render_template('items.html', items=items)

@app.route('/add-item', methods=['POST'])
def add_item():
    item = ItemTemplate(
        name=request.form.get('name'),
        description=request.form.get('description'),
        default_price=float(request.form.get('default_price')),
        unit=request.form.get('unit', 'SQFT')
    )
    db.session.add(item)
    db.session.commit()
    flash('Item template added successfully!', 'success')
    return redirect(url_for('items'))

@app.route('/delete-item/<int:item_id>')
def delete_item(item_id):
    item = ItemTemplate.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('Item template deleted successfully!', 'success')
    return redirect(url_for('items'))

@app.route('/api/item-templates')
def api_item_templates():
    items = ItemTemplate.query.all()
    return jsonify([{'id': i.id, 'name': i.name, 'description': i.description, 'default_price': i.default_price, 'unit': i.unit} for i in items])

@app.route('/api/search-contacts')
def search_contacts():
    query = request.args.get('query', '')
    if query:
        contacts = Contact.query.filter(Contact.name.ilike(f'%{query}%')).limit(10).all()
    else:
        contacts = []
    return jsonify([{'id': c.id, 'name': c.name, 'address': c.address, 'phone': c.phone, 'email': c.email} for c in contacts])

@app.route('/quote/<int:quote_id>/send-email', methods=['GET', 'POST'])
def send_email(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    
    if request.method == 'POST':
        to_email = request.form.get('to_email')
        subject = request.form.get('subject')
        message_body = request.form.get('message')
        attached_file = request.files.get('attachment')
        
        if not to_email:
            flash('Please enter recipient email address', 'danger')
            return redirect(url_for('send_email', quote_id=quote_id))
        if not attached_file or attached_file.filename == '':
            flash('Please attach a PDF file', 'danger')
            return redirect(url_for('send_email', quote_id=quote_id))
        if not attached_file.filename.endswith('.pdf'):
            flash('Please attach a valid PDF file', 'danger')
            return redirect(url_for('send_email', quote_id=quote_id))
        
        try:
            msg = Message(subject=subject, recipients=[to_email], body=message_body, sender=app.config['MAIL_USERNAME'])
            msg.attach(attached_file.filename, 'application/pdf', attached_file.read())
            mail.send(msg)
            flash(f'✅ Email sent successfully to {to_email}!', 'success')
            return redirect(url_for('view_quote', quote_id=quote.id))
        except Exception as e:
            flash(f'❌ Error: {str(e)}', 'danger')
            return redirect(url_for('send_email', quote_id=quote_id))
    
    return render_template('email_composer.html', quote=quote)

@app.route('/quote/<int:quote_id>/print-view')
def quote_print_view(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    quote.total_amount = float(quote.total_amount or 0)
    quote.cleaning_amount = float(quote.cleaning_amount or 0)
    quote.discount_amount = float(quote.discount_amount or 0)
    quote.discount_value = float(quote.discount_value or 0)
    quote.apply_gst = bool(quote.apply_gst) if quote.apply_gst is not None else True
    quote.cleaning_charges = bool(quote.cleaning_charges) if quote.cleaning_charges is not None else False
    return render_template('quote_template.html', quote=quote)

@app.route('/download-pdf/<quote_no>')
def download_pdf(quote_no):
    quote = Quote.query.filter_by(quote_no=quote_no).first_or_404()
    html_content = render_template('quote_pdf.html', quote=quote)
    pdf_file = io.BytesIO()
    pisa.CreatePDF(html_content, dest=pdf_file)
    pdf_file.seek(0)
    return send_file(pdf_file, as_attachment=True, download_name=f'Estimate_{quote.quote_no}.pdf', mimetype='application/pdf')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
